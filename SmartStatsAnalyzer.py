# Smart Stats Analyzer - Final Version with All Enhancements
import customtkinter as ctk
from tkinter import filedialog, messagebox
from PIL import Image, ImageTk
import pandas as pd
from statistics import mean, median, mode, variance, stdev
import matplotlib.pyplot as plt
import io, os, json
from fpdf import FPDF
import openpyxl
from scipy.stats import shapiro

# Globals
df = None
current_data = []
last_chart_path = "last_chart.png"

# UI setup
ctk.set_appearance_mode("light")
ctk.set_default_color_theme("blue")

splash = ctk.CTk()
splash.geometry("500x300")
splash.title("Welcome")
ctk.CTkLabel(splash, text="\U0001F4CA Statistical Analyzer", font=("Segoe UI", 26, "bold")).pack(pady=20)
ctk.CTkLabel(splash, text="Developed by Abdulrahman Hamdi", font=("Segoe UI", 16)).pack(pady=10)
ctk.CTkLabel(splash, text="Initializing...", font=("Segoe UI", 14)).pack(pady=40)
splash.after(2000, splash.destroy)
splash.mainloop()

app = ctk.CTk()
app.geometry("880x760")
app.title("Smart Stats Analyzer")

# Dark mode toggle
def toggle_theme():
    current = ctk.get_appearance_mode()
    new_mode = "dark" if current == "light" else "light"
    ctk.set_appearance_mode(new_mode)

theme_button = ctk.CTkButton(app, text="\U0001F319 Toggle Theme", command=toggle_theme, width=140)
theme_button.place(x=700, y=20)

# Tabs
tabs = ctk.CTkTabview(app)
tabs.pack(padx=20, pady=20, fill="both", expand=True)

main_tab = tabs.add("\U0001F4CA Analysis")
chart_tab = tabs.add("\U0001F4C8 Graph")
about_tab = tabs.add("\u2139\ufe0f About")

ctk.CTkLabel(main_tab, text="Enter comma-separated numbers or load from CSV:").pack(pady=10)
entry = ctk.CTkEntry(main_tab, width=700, height=40, font=("Segoe UI", 14))
entry.pack(pady=5)
entry.bind("<Return>", lambda event: manual_calculate())

# Buttons
btn_frame = ctk.CTkFrame(main_tab, fg_color="transparent")
btn_frame.pack(pady=10)
ctk.CTkButton(btn_frame, text="Calculate", width=120, command=lambda: manual_calculate()).grid(row=0, column=0, padx=8)
ctk.CTkButton(btn_frame, text="Load CSV", width=120, command=lambda: load_csv()).grid(row=0, column=1, padx=8)
ctk.CTkButton(btn_frame, text="Show Graphs", width=120, command=lambda: show_graph()).grid(row=0, column=2, padx=8)
ctk.CTkButton(btn_frame, text="Export to PDF", width=120, command=lambda: export_pdf()).grid(row=0, column=3, padx=8)
ctk.CTkButton(btn_frame, text="Export to Excel", width=120, command=lambda: export_excel()).grid(row=0, column=4, padx=8)
ctk.CTkButton(btn_frame, text="Clear", width=120, command=lambda: clear_entry()).grid(row=0, column=5, padx=8)
ctk.CTkButton(btn_frame, text="Reset All", width=120, command=lambda: reset_all()).grid(row=0, column=6, padx=8)

combo = ctk.CTkOptionMenu(main_tab, values=["Select a column"], command=lambda choice: select_column(choice))
combo.pack(pady=5)

# Add manual_calculate function
def manual_calculate():
    try:
        values = entry.get().replace(" ", "").split(",")
        numbers = list(map(float, values))
        update_result_cards(numbers)
        global current_data
        current_data = numbers
    except:
        messagebox.showerror("Error", "Please enter valid comma-separated numbers.")

card_container = ctk.CTkFrame(main_tab, corner_radius=12)
card_container.pack(pady=15, padx=10, fill="both", expand=True)

card_tendency = ctk.CTkFrame(card_container, corner_radius=10, fg_color="#e6f2ff")
card_dispersion = ctk.CTkFrame(card_container, corner_radius=10, fg_color="#fef9e7")
card_freq = ctk.CTkFrame(card_container, corner_radius=10, fg_color="#f9ebea")

card_tendency.pack(pady=10, padx=15, fill="x")
card_dispersion.pack(pady=10, padx=15, fill="x")
card_freq.pack(pady=10, padx=15, fill="x")

lbl_tendency = ctk.CTkLabel(card_tendency, text="\U0001F4CC Central Tendency", font=("Segoe UI", 16, "bold"))
lbl_dispersion = ctk.CTkLabel(card_dispersion, text="\U0001F4C9 Dispersion", font=("Segoe UI", 16, "bold"))
lbl_freq = ctk.CTkLabel(card_freq, text="\U0001F4E6 Frequencies", font=("Segoe UI", 16, "bold"))

lbl_tendency.pack(anchor="w", padx=10, pady=(5, 0))
lbl_dispersion.pack(anchor="w", padx=10, pady=(5, 0))
lbl_freq.pack(anchor="w", padx=10, pady=(5, 0))

lbl_mean = ctk.CTkLabel(card_tendency, text="")
lbl_median = ctk.CTkLabel(card_tendency, text="")
lbl_mode = ctk.CTkLabel(card_tendency, text="")

lbl_variance = ctk.CTkLabel(card_dispersion, text="")
lbl_std = ctk.CTkLabel(card_dispersion, text="")
lbl_range = ctk.CTkLabel(card_dispersion, text="")
lbl_shapiro = ctk.CTkLabel(card_dispersion, text="")

lbl_mean.pack(anchor="w", padx=15)
lbl_median.pack(anchor="w", padx=15)
lbl_mode.pack(anchor="w", padx=15)
lbl_variance.pack(anchor="w", padx=15)
lbl_std.pack(anchor="w", padx=15)
lbl_range.pack(anchor="w", padx=15)
lbl_shapiro.pack(anchor="w", padx=15)

freq_frame = ctk.CTkFrame(card_freq)
freq_frame.pack(padx=15, pady=5, fill="x")

# Chart tab
chart_label = ctk.CTkLabel(chart_tab, text="Graphs will appear here after clicking 'Show Graphs'", font=("Segoe UI", 14))
chart_label.pack(pady=10)

btn_chart_frame = ctk.CTkFrame(chart_tab, fg_color="transparent")
btn_chart_frame.pack(pady=10)
ctk.CTkButton(btn_chart_frame, text="Histogram", command=lambda: show_graph()).grid(row=0, column=0, padx=10)
ctk.CTkButton(btn_chart_frame, text="Boxplot", command=lambda: draw_boxplot()).grid(row=0, column=1, padx=10)
ctk.CTkButton(btn_chart_frame, text="Pie Chart", command=lambda: draw_piechart()).grid(row=0, column=2, padx=10)


# About tab
ctk.CTkLabel(about_tab, text="\U0001F4CA Statistical Analyzer App\nVersion 1.0\nDeveloped by Abdulrahman Hamdi\n\u00a9 2025", font=("Segoe UI", 15)).pack(pady=40)

# Core functions
def update_result_cards(numbers):
    avg, med, mod_ = mean(numbers), median(numbers), mode(numbers)
    var, std, rng = variance(numbers), stdev(numbers), max(numbers) - min(numbers)
    lbl_mean.configure(text=f"\U0001F4CC Mean: {avg:.4f}")
    lbl_median.configure(text=f"\U0001F3AF Median: {med:.4f}")
    lbl_mode.configure(text=f"\U0001F3B2 Mode: {mod_:.4f}")
    lbl_variance.configure(text=f"\U0001F300 Variance: {var:.4f}")
    lbl_std.configure(text=f"\U0001F4CF Std Dev: {std:.4f}")
    lbl_range.configure(text=f"\U0001F4D0 Range: {rng:.4f}")
    try:
        stat, p = shapiro(numbers)
        result = "\u2714\ufe0f Normally Distributed" if p > 0.05 else "\u274C Not Normal"
        lbl_shapiro.configure(text=f"\U0001F4CA Shapiro-Wilk: p={p:.4f} → {result}")
    except:
        lbl_shapiro.configure(text="\U0001F4CA Shapiro-Wilk: Error")
    for widget in freq_frame.winfo_children(): widget.destroy()
    freqs = {}
    for val in numbers: freqs[val] = freqs.get(val, 0) + 1
    for k, v in freqs.items():
        ctk.CTkLabel(freq_frame, text=f"{k}: {v} times").pack(anchor="w")
    with open("session.json", "w") as f:
        json.dump(numbers, f)

def clear_entry(): entry.delete(0, "end")

def reset_all():
    entry.delete(0, "end")
    combo.set("Select a column")
    global current_data
    current_data = []
    for lbl in [lbl_mean, lbl_median, lbl_mode, lbl_variance, lbl_std, lbl_range, lbl_shapiro]: lbl.configure(text="")
    for widget in freq_frame.winfo_children(): widget.destroy()
    chart_label.configure(text="Graphs will appear here after clicking 'Show Graphs'", image=None)
    chart_label.image = None

def load_csv():
    global df
    file = filedialog.askopenfilename(filetypes=[
        ("CSV Files", "*.csv"),
        ("Excel Files", "*.xls;*.xlsx"),
        ("JSON Files", "*.json"),
        ("Text Files", "*.txt")
    ])
    if not file: return
    try:
        if file.endswith(".csv"): df = pd.read_csv(file)
        elif file.endswith(".xlsx") or file.endswith(".xls"): df = pd.read_excel(file)
        elif file.endswith(".json"): df = pd.read_json(file)
        elif file.endswith(".txt"): df = pd.read_csv(file, delimiter="\t")
        else: raise ValueError("Unsupported file type.")
        cols = df.columns.tolist()
        combo.configure(values=cols)
        combo.set("Select a column")
        messagebox.showinfo("Loaded", "File loaded. Please choose a column.")
    except Exception as e:
        messagebox.showerror("Error", str(e))

def select_column(choice):
    global current_data
    try:
        values = df[choice].dropna().astype(float).tolist()
        entry.delete(0, "end")
        entry.insert(0, ", ".join(map(str, values)))
        update_result_cards(values)
        current_data = values
    except Exception as e:
        messagebox.showerror("Error", str(e))

def show_graph():
    if not current_data:
        messagebox.showinfo("No Data", "Please enter data or select column first.")
        return
    fig, ax = plt.subplots()
    ax.hist(current_data, bins=10)
    ax.set_title("Histogram")
    ax.set_xlabel("Values")
    ax.set_ylabel("Frequency")
    buf = io.BytesIO()
    fig.savefig(buf, format="png")
    buf.seek(0)
    plt.close()
    img = Image.open(buf)
    img = img.resize((600, 400))
    img.save(last_chart_path)
    photo = ImageTk.PhotoImage(img)
    chart_label.configure(image=photo, text="")
    chart_label.image = photo

def draw_boxplot():
    if not current_data: return
    fig, ax = plt.subplots()
    ax.boxplot(current_data, vert=False)
    ax.set_title("Boxplot")
    ax.set_xlabel("Values")
    buf = io.BytesIO()
    fig.savefig(buf, format="png")
    buf.seek(0)
    plt.close()
    img = Image.open(buf)
    img = img.resize((600, 400))
    img.save(last_chart_path)
    photo = ImageTk.PhotoImage(img)
    chart_label.configure(image=photo, text="")
    chart_label.image = photo

def draw_piechart():
    if not current_data: return
    freqs = {}
    for val in current_data: freqs[val] = freqs.get(val, 0) + 1
    labels = [str(k) for k in freqs.keys()]
    sizes = list(freqs.values())
    fig, ax = plt.subplots()
    ax.pie(sizes, labels=labels, autopct='%1.1f%%', startangle=90)
    ax.axis("equal")
    ax.set_title("Pie Chart (Frequencies)")
    buf = io.BytesIO()
    fig.savefig(buf, format="png")
    buf.seek(0)
    plt.close()
    img = Image.open(buf)
    img = img.resize((600, 400))
    img.save(last_chart_path)
    photo = ImageTk.PhotoImage(img)
    chart_label.configure(image=photo, text="")
    chart_label.image = photo

def export_pdf():
    if not current_data: return
    try:
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Arial", size=12)
        pdf.cell(200, 10, txt="Statistical Report", ln=True, align="C")
        pdf.ln()
        pdf.cell(200, 10, txt=f"Mean: {mean(current_data):.4f}", ln=True)
        pdf.cell(200, 10, txt=f"Median: {median(current_data):.4f}", ln=True)
        pdf.cell(200, 10, txt=f"Mode: {mode(current_data):.4f}", ln=True)
        pdf.cell(200, 10, txt=f"Variance: {variance(current_data):.4f}", ln=True)
        pdf.cell(200, 10, txt=f"Std Dev: {stdev(current_data):.4f}", ln=True)
        pdf.cell(200, 10, txt=f"Range: {max(current_data)-min(current_data):.4f}", ln=True)
        pdf.ln(10)
        pdf.cell(200, 10, txt="Frequencies:", ln=True)
        freqs = {}
        for val in current_data:
            freqs[val] = freqs.get(val, 0) + 1
        for k, v in freqs.items():
            pdf.cell(200, 10, txt=f"{k}: {v} times", ln=True)
        if os.path.exists(last_chart_path):
            try:
                pdf.image(last_chart_path, x=10, y=None, w=180)
            except: pass
        filepath = filedialog.asksaveasfilename(defaultextension=".pdf", filetypes=[("PDF files", "*.pdf")])
        if filepath:
            pdf.output(filepath)
            messagebox.showinfo("Success", f"PDF saved to {filepath}")
    except Exception as e:
        messagebox.showerror("PDF Error", str(e))

def export_excel():
    if not current_data: return
    try:
        df_export = pd.DataFrame({"Values": current_data})
        freqs = df_export["Values"].value_counts().reset_index()
        freqs.columns = ["Value", "Frequency"]
        filepath = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if filepath:
            with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
                df_export.to_excel(writer, index=False, sheet_name="Data")
                freqs.to_excel(writer, index=False, sheet_name="Frequencies")
            messagebox.showinfo("Success", f"Excel file saved to {filepath}")
    except Exception as e:
        messagebox.showerror("Excel Error", str(e))

def load_session():
    global current_data
    try:
        with open("session.json", "r") as f:
            current_data = json.load(f)
            entry.insert(0, ", ".join(map(str, current_data)))
            update_result_cards(current_data)
    except: pass

load_session()
app.mainloop()
